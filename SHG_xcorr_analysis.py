#!/usr/bin/env python3
"""
Written by Mingzhen Yu with assitance of Claude Fable5.
Cross-correlation analysis and Figure 9 for sediment-hosted glass (SHG)
geochemical time series vs. local relative sea level, southern Reykjanes Ridge.

This single script reproduces every statistic in Section 4 of the manuscript and
draws the Figure 9. It performs:

  1. RSL vs dRSL lead check: the offset (kyr) at which the rate of sea-level
     change best correlates with sea level itself, per analysis window.
  2. Lagged Pearson cross-correlation between the unsmoothed 2-kyr binned
     geochemical series (La, La/Sm, Rb, Ba, Th, K2O; cores GeoB25444-1 "core 44"
     and GeoB25422-1 "core 22") and the local sea-level record (RSL and its rate
     dRSL; Reilly et al., 2025), for lag windows +/-14 kyr (primary) and
     +/-20 kyr (sensitivity).
  3. Significance against two surrogate nulls (NSUR each), with the maximum
     correlation across all tested lags as the test statistic (accounting for
     the lag search; one-sided, positive correlations):
       - AR(1) surrogates matched to each series' lag-1 autocorrelation and
         variance (liberal null);
       - phase-randomized surrogates preserving each series' full power
         spectrum (conservative null).
  4. Five-bin (10 kyr) smoothed comparison (visualization reference only).
  5. Figure 9: 2x2 panels of r(lag) curves (core x target) with peak markers
     encoding significance: filled = p < 0.05 vs phase-randomized null;
     half-filled = p < 0.05 vs AR(1) null only; open = neither.

Conventions:
  - Positive lag = geochemistry lags the sea-level forcing (geochemistry at age
    a is compared with the sea-level series at older age a + lag).
  - Core 22 restricted to <= 115 ka; unsampled bins (13, 17, 93 ka) filled by
    linear interpolation. Lags with < 20% overlap excluded. dRSL > 0 = rise.

Input : Excel workbook with sheets 'core 22', 'core 44' (Age + element columns)
        and 'relative sea level' (Age_ka, RSL_m, dRSL_m/kyr).
Output: shg_xcorr_results.csv  (all windows/cores/targets/elements: r, lag,
                                p-values, edge flag, smoothed comparison)
        fig9_curves.csv        (full r(lag) curves, +/-14 kyr, for replotting)
        fig9_significance.csv  (peak r, lag, p-values, significance class)
        fig9.png/.pdf    (the figure)

Reproducibility: the random generator is re-seeded per (core, target, element)
combination, so all p-values are exactly reproducible and identical between the
results table and the figure. Monte Carlo error at NSUR = 2000 is ~+/-0.005-0.01
on p; raise NSUR (e.g., 10000) for final manuscript values.

Usage: python3 SHG_xcorr_analysis.py [data.xlsx]     (default 'SHG_data_for_xcorr.xlsx')
                                                     (or 'SHG_data_for_xcorr_fine_tuned.xlsx')
Dependencies: numpy, openpyxl, matplotlib
"""
import sys
import csv
import numpy as np
import openpyxl
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

# ---------------------------- configuration --------------------------------
DATA = sys.argv[1] if len(sys.argv) > 1 else 'SHG_data_for_xcorr.xlsx'
NSUR = 10000                               # ±0.002 at p = 0.05
SEED = 42
DT = 2.0                                   # bin width, kyr
WINDOWS = [20,14]                         # lag windows, kyr (first = primary, used for figure)
ELEMENTS = ['La', 'La/Sm', 'Rb', 'Ba', 'Th', 'K2O']
CORES = [('core 44', 74), ('core 22', 115)]  # (sheet, max age used, ka)
MIN_OVERLAP_FRAC = 0.20
COLORS = {'La': "peru", 'La/Sm': "darkcyan", 'Rb': "firebrick",
          'Ba': "slateblue", 'Th': "olivedrab", 'K2O': "palevioletred"}

# ------------------------------ helpers ------------------------------------
def load(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for name in wb.sheetnames:
        rows = list(wb[name].iter_rows(values_only=True))
        hdr = rows[0]
        body = [r for r in rows[1:] if isinstance(r[0], (int, float))]
        arr = np.array(body, dtype=float)
        out[name] = {h: arr[:, i] for i, h in enumerate(hdr) if h}
    return out


def bin_avg(centers, xa, xv, half=DT / 2):
    """Average xv within [center-half, center+half) for each bin center."""
    return np.array([np.nanmean(xv[(xa >= a - half) & (xa < a + half)]) for a in centers])


def interp_nan(x):
    x = x.copy()
    n = np.isnan(x)
    if n.any():
        x[n] = np.interp(np.flatnonzero(n), np.flatnonzero(~n), x[~n])
    return x


def smooth5(x):
    """Five-bin centered moving average (edge-truncated)."""
    return np.array([np.nanmean(x[max(0, i - 2):i + 3]) for i in range(len(x))])


def rmat(G, s):
    """Pearson r between each row of matrix G and vector s."""
    G = G - G.mean(1, keepdims=True)
    s = s - s.mean()
    return (G @ s) / (np.sqrt((G ** 2).sum(1)) * np.sqrt((s ** 2).sum()) + 1e-300)


def seg(G, s, k):
    """Overlapping segments of G (2-D) and s for lag step k (positive = G lags s)."""
    N = G.shape[1]
    if k > 0:
        return G[:, :N - k], s[k:]
    if k < 0:
        return G[:, -k:], s[:N + k]
    return G, s


def r_curve(g, s, ks, min_n):
    """r at each lag step for a single series; NaN where overlap too small."""
    out = np.full(len(ks), np.nan)
    for i, k in enumerate(ks):
        gg, ss = seg(g[None, :], s, k)
        if gg.shape[1] >= max(min_n, 5):
            out[i] = rmat(gg, ss)[0]
    return out


def xcorr_max(G, s, ks, min_n):
    """Maximum r over lag steps ks for each row of G (the test statistic)."""
    best = np.full(G.shape[0], -np.inf)
    for k in ks:
        gg, ss = seg(G, s, k)
        if gg.shape[1] < max(min_n, 5):
            continue
        best = np.maximum(best, rmat(gg, ss))
    return best


def phase_rand(x, n_sur, rng):
    """Surrogates preserving the full power spectrum of x (phases randomized)."""
    n = len(x)
    X = np.fft.rfft(x - x.mean())
    ph = rng.uniform(0, 2 * np.pi, (n_sur, len(X)))
    ph[:, 0] = 0
    if n % 2 == 0:
        ph[:, -1] = 0
    return np.fft.irfft(np.abs(X) * np.exp(1j * ph), n, axis=1)


def ar1(x, n_sur, rng):
    """AR(1) surrogates matched to lag-1 autocorrelation and variance of x."""
    x0 = x - x.mean()
    r1 = np.clip(np.corrcoef(x0[:-1], x0[1:])[0, 1], -0.99, 0.99)
    n = len(x)
    e = rng.normal(size=(n_sur, n))
    y = np.zeros((n_sur, n))
    y[:, 0] = e[:, 0]
    for i in range(1, n):
        y[:, i] = r1 * y[:, i - 1] + np.sqrt(1 - r1 ** 2) * e[:, i]
    return y * x0.std()


# ------------------------------- main --------------------------------------
def main():
    d = load(DATA)
    sl = d['relative sea level']
    o = np.argsort(sl['Age_ka'])
    sa, sr, sd = sl['Age_ka'][o], sl['RSL_m'][o], sl['dRSL_m/kyr'][o]
    m = ~np.isnan(sd)

    # ---- 1. RSL vs dRSL lead check ----------------------------------------
    print('RSL vs dRSL lead check (offset of max correlation):')
    for sheet, maxage in CORES:
        ac = np.arange(1, maxage + 1, DT)
        rsl, drsl = bin_avg(ac, sa, sr), bin_avg(ac, sa[m], sd[m])
        curve = [(2 * k, rmat(*seg(rsl[None, :], drsl, k))[0]) for k in range(0, 11)]
        lead, r = max(curve, key=lambda t: t[1])
        print(f'  window 0-{maxage} ka ({sheet}): dRSL leads RSL by {lead} kyr (r = {r:.2f})')

    # ---- 2-4. cross-correlation, significance, smoothed comparison --------
    results = []          # all rows for shg_xcorr_results.csv
    fig_data = {}         # (core, target, element) -> dict, primary window, for the figure
    W_PRIMARY = WINDOWS[0]
    print(f"\n{'window':<8}{'core':<9}{'target':<7}{'elem':<7}{'r':>6}{'lag':>5}"
          f"{'p_phase':>9}{'p_AR1':>7}{'edge':>6} | {'r_5bin':>7}{'lag5':>5}")
    for W in WINDOWS:
        ks = list(range(-W // 2, W // 2 + 1))
        lags = np.array(ks) * DT
        for sheet, maxage in CORES:
            ages = d[sheet]['Age']
            sel = ages <= maxage
            ac = ages[sel]
            assert not np.any(np.diff(ac) != DT), f'{sheet}: age bins not contiguous'
            targets = {'RSL': bin_avg(ac, sa, sr), 'dRSL': bin_avg(ac, sa[m], sd[m])}
            for tname, t in targets.items():
                assert not np.any(np.isnan(t))
                for el in ELEMENTS:
                    g = interp_nan(d[sheet][el][sel])
                    min_n = int(MIN_OVERLAP_FRAC * len(g))
                    rc = r_curve(g, t, ks, min_n)
                    ipk = int(np.nanargmax(rc))
                    rpk, lagpk = float(rc[ipk]), int(lags[ipk])
                    rng = np.random.default_rng(SEED)   # per-combination seed
                    n_ph = xcorr_max(phase_rand(g, NSUR, rng), t, ks, min_n)
                    n_a = xcorr_max(ar1(g, NSUR, rng), t, ks, min_n)
                    p_ph = (np.sum(n_ph >= rpk) + 1) / (NSUR + 1)
                    p_a = (np.sum(n_a >= rpk) + 1) / (NSUR + 1)
                    sig = 'phase' if p_ph < 0.05 else 'ar1' if p_a < 0.05 else 'none'
                    edge = abs(lagpk) == W
                    g5, t5 = smooth5(g), smooth5(t)
                    rc5 = r_curve(g5, t5, ks, min_n)
                    i5 = int(np.nanargmax(rc5))
                    print(f'±{W:<7}{sheet:<9}{tname:<7}{el:<7}{rpk:>6.2f}{lagpk:>5}'
                          f'{p_ph:>9.3f}{p_a:>7.3f}{"EDGE" if edge else "":>6}'
                          f' | {rc5[i5]:>7.2f}{int(lags[i5]):>5}')
                    results.append(dict(window=f'±{W}', core=sheet, target=tname,
                                        element=el, r=round(rpk, 3), lag_kyr=lagpk,
                                        p_phase=round(p_ph, 4), p_ar1=round(p_a, 4),
                                        edge=edge, significance=sig,
                                        r_5bin=round(float(rc5[i5]), 3),
                                        lag_5bin_kyr=int(lags[i5])))
                    if W == W_PRIMARY:
                        fig_data[(sheet, tname, el)] = dict(
                            lags=lags, rc=rc, rpk=rpk, lagpk=lagpk,
                            p_ph=p_ph, p_a=p_a, sig=sig)

    with open('shg_xcorr_results.csv', 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=results[0].keys())
        w.writeheader()
        w.writerows(results)

    # ---- 5. Figure 9 (primary window) -------------------------------------
    curve_rows, sig_rows = [], []
    fig, axes = plt.subplots(2, 2, figsize=(7.2, 6.2), dpi=300, sharex='col')
    panels = [('core 44', 'dRSL', axes[0, 0], 'a'), ('core 44', 'RSL', axes[0, 1], 'b'),
              ('core 22', 'dRSL', axes[1, 0], 'c'), ('core 22', 'RSL', axes[1, 1], 'd')]
    for sheet, tname, ax, lab in panels:
        for el in ELEMENTS:
            fd = fig_data[(sheet, tname, el)]
            c = COLORS[el]
            if el == 'K2O':
                ax.plot(fd['lags'], fd['rc'], 's-', color=c, lw=1.1, mec='k', mew=0.5, ms=3, label='K$_2$O')
            else:
                ax.plot(fd['lags'], fd['rc'], 's-', color=c, lw=1.1, mec='k', mew=0.5, ms=3, label=el)
            style = dict(ms=7, mec=c, mew=1.2, zorder=5)
            if fd['sig'] == 'phase':
                ax.plot(fd['lagpk'], fd['rpk'], 'o', mfc=c, **style)
            elif fd['sig'] == 'ar1':
                ax.plot(fd['lagpk'], fd['rpk'], 'o', mfc=c, fillstyle='left',
                        markerfacecoloralt='white', **style)
            else:
                ax.plot(fd['lagpk'], fd['rpk'], 'o', mfc='white', **style)
            ax.axvline(fd['lagpk'], color=c, lw=0.5, ls='--', alpha=0.5)
            for L, r in zip(fd['lags'], fd['rc']):
                curve_rows.append(dict(core=sheet, target=tname, element=el,
                                       lag_kyr=int(L), r=round(float(r), 4)))
            sig_rows.append(dict(core=sheet, target=tname, element=el,
                                 r_peak=round(fd['rpk'], 3), lag_peak_kyr=fd['lagpk'],
                                 p_phase=round(fd['p_ph'], 4), p_ar1=round(fd['p_a'], 4),
                                 significance=fd['sig']))
        ax.axhline(0, color='k', lw=0.5,alpha=0.8)
        ax.set_title(f'{sheet} vs. {tname}', loc='left', fontsize=10, fontfamily='Arial')
    for ax in axes[1,:]:
        ax.set_xlabel('Lead/Lag (kyr)', fontsize=10, fontfamily='Arial')
    for ax in axes[:, 0]:
        ax.set_ylabel('Correlation Coefficient', fontsize=10, fontfamily='Arial')
        ax.set_xlim(-3,21)
        ax.set_xticks(np.arange(-2,21,2))
    for ax in axes[:,1]:
        ax.set_xlim(-15,15)
        ax.set_xticks(np.arange(-14,15,2))
    for ax in axes.flat:
        ax.tick_params(axis='both', which='major', labelsize=10, length=3, width=0.5)
        for label in ax.get_xticklabels() + ax.get_yticklabels(): 
            label.set_fontfamily('Arial')
    h, l = axes[0, 0].get_legend_handles_labels()
    leg = axes[0, 0].legend(h, l, frameon=True, ncol=2, loc='lower right',
                      bbox_to_anchor=(1.01,-0.01),
                      prop={'family': 'Arial','size': 10},
                      columnspacing=0.5, borderpad=0.1, handlelength=1.0,
                      handletextpad=0.1, edgecolor='black', fancybox=False,
                      labelspacing=0.1, markerscale=0.75)
    leg.get_frame().set_linewidth(0.5)
    # fig.text(0.01, 0.005,
    #          'peak markers: filled = p<0.05 (phase-randomized null); '
    #          'half-filled = p<0.05 (AR(1) only); open = n.s.', fontsize=6.5)
    axes[0,0].text(-0.14, 0.97, 'a', transform=axes[0,0].transAxes, fontsize=10, fontweight='bold', fontname='Arial')
    axes[0,1].text(-0.14, 0.97, 'b', transform=axes[0,1].transAxes, fontsize=10, fontweight='bold', fontname='Arial')
    axes[1,0].text(-0.14, 0.97, 'c', transform=axes[1,0].transAxes, fontsize=10, fontweight='bold', fontname='Arial')
    axes[1,1].text(-0.14, 0.97, 'd', transform=axes[1,1].transAxes, fontsize=10, fontweight='bold', fontname='Arial')
    plt.tight_layout()
    # plt.savefig('fig9_draft.png', dpi=300)
    plt.savefig('fig9_draft.pdf')

    with open('fig9_curves.csv', 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=curve_rows[0].keys())
        w.writeheader()
        w.writerows(curve_rows)
    with open('fig9_significance.csv', 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=sig_rows[0].keys())
        w.writeheader()
        w.writerows(sig_rows)
    print('\nSaved: shg_xcorr_results.csv, fig9_curves.csv, fig9_significance.csv, '
          'fig9_draft.png/.pdf')


if __name__ == '__main__':
    main()